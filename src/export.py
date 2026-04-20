import contextlib
import json
import logging
import os
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.database import Database

logger = logging.getLogger(__name__)

SEVERITY_COLORS = {
    1: "C6EFCE",  # green
    2: "D9E1F2",  # light blue
    3: "FCE4D6",  # light orange
    4: "F4CCCC",  # light red
    5: "EA9999",  # red
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_to_xlsx(db: Database, output_path: str | None = None) -> str:
    if output_path is None:
        os.makedirs("data/export", exist_ok=True)
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        output_path = f"data/export/ci_monitor_{timestamp}.xlsx"

    wb = Workbook()

    _write_signals_sheet(wb, db)
    _write_competitors_sheet(wb, db)
    _write_summary_sheet(wb, db)

    # Remove default empty sheet if others exist
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(output_path)
    logger.info("Exported to %s", output_path)
    return output_path


def _style_header(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, max_width: int = 50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_signals_sheet(wb: Workbook, db: Database):
    ws = wb.active
    ws.title = "Signals"

    headers = [
        "ID", "Detected At", "Competitor", "Source", "Type",
        "Title", "Content", "URL", "Severity", "Tags",
        "Published At", "Change Summary",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    signals = db.get_signals(limit=0)
    for sig in signals:
        tags = sig.get("tags", "")
        if tags:
            with contextlib.suppress(json.JSONDecodeError, TypeError):
                tags = ", ".join(json.loads(tags))

        row = [
            sig["id"][:8],
            sig["detected_at"],
            sig["competitor_id"],
            sig["source"],
            sig["signal_type"],
            sig["title"],
            (sig.get("content") or "")[:500],
            sig.get("url", ""),
            sig["severity"],
            tags,
            sig.get("published_at", ""),
            sig.get("change_summary", ""),
        ]
        ws.append(row)

        # Color row by severity
        row_num = ws.max_row
        severity = sig["severity"]
        if severity in SEVERITY_COLORS:
            fill = PatternFill(
                start_color=SEVERITY_COLORS[severity],
                end_color=SEVERITY_COLORS[severity],
                fill_type="solid",
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).border = THIN_BORDER

    # Add autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _auto_width(ws)

    # Freeze header row
    ws.freeze_panes = "A2"


def _write_competitors_sheet(wb: Workbook, db: Database):
    ws = wb.create_sheet("Competitors")

    headers = ["ID", "Name", "Parent Group", "Tier", "IČO", "Config Path"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for comp in db.get_competitors():
        ws.append([
            comp["id"],
            comp["name"],
            comp.get("parent_group", ""),
            comp["tier"],
            comp.get("ico", ""),
            comp.get("config_path", ""),
        ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _write_summary_sheet(wb: Workbook, db: Database):
    ws = wb.create_sheet("Summary")

    headers = ["Competitor", "Total Signals", "Avg Severity", "Last Signal", "Sources"]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Aggregate per competitor
    rows = db.conn.execute(
        """SELECT competitor_id,
                  COUNT(*) as total,
                  ROUND(AVG(severity), 1) as avg_sev,
                  MAX(detected_at) as last_signal,
                  GROUP_CONCAT(DISTINCT source) as sources
           FROM signals
           GROUP BY competitor_id
           ORDER BY total DESC"""
    ).fetchall()

    for r in rows:
        ws.append([
            r["competitor_id"],
            r["total"],
            r["avg_sev"],
            r["last_signal"],
            r["sources"],
        ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _auto_width(ws)
    ws.freeze_panes = "A2"
